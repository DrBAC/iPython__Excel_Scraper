#######################################################################################################
#######################################################################################################
#           this Python3 Program was written by Ben Coombs for use probing excel files                #
#              at CPI NetPark - Printable Electronics. It is not for general use                      #
#          and should be considered to be in beta-test mode. BAC Copyright March 2016.                #
#######################################################################################################
###########           ##############           #############         #################        #########

# new version featuring AS-YOU-GO WRITE OUT
# new version featuring LOCAL ORIGINS FOR BETTER SEARCHING
# searches for VTH, IONOFF and VTO locally now (directly indexed)
# searches now performed using REGEX
# formatting of output is now correct - incorporated in the function definition

##############          STILL NEED CATCHALLS FOR WHEN THE PARAMETER ISN'T FOUND        ################

import openpyxl
from openpyxl import load_workbook
from openpyxl.cell import get_column_letter, column_index_from_string
import time
import os
import pandas as pd
import csv
import re

#start clock to calculate timings
start_time = time.clock()
      
# initiate some variables/lists - substrate is index so can count number of files searched thru	  
substrate = 0                                               
outputfile = []
channel_data = []
substrate_data = []

print('Excel File Name',',','CL',',','CW',',','Mobility',',','SDev',',','On/Off'
      ,',','VTO',',','VTH',',','Yield')

# file to write to is opened, named based on date
date = time.strftime('DIRECTORY TEST DATA OVERVIEW %Y-%m-%d.csv')#-%H%M%S)
with open(date, "w", newline='') as f:
    writer = csv.writer(f)
    writer.writerow(["File", "Date of test", "CL", "CW", "Mobility", "SDev", "On/Off", 
                     "VTO", "VTH", "Yield"])

    for root, dirs, files in os.walk(".", topdown=True):
        for fname in files:
            if fname.endswith('.xlsm') or fname.endswith('.xlsx'):
                wb = load_workbook(os.path.join(root, fname), data_only=True)
                names = wb.get_sheet_names()

                if 'validated Summary Data' in names:
                    ws = wb.get_sheet_by_name('validated Summary Data')

                    substrate = substrate + 1

                    Row_limit = ws.max_row
                    Column_limit = ws.max_column
                    Col_limit = get_column_letter(Column_limit)

                    def SearchXLWithRE(regex):
                        for i in range(1, Row_limit):
                            for j in range(1, Column_limit):
                                query = ws.cell(row = i, column = j).value
                                if type(query) == str:
                                    test = regex.search(query)
                                    if test: # checks if regex has found a match
                                        return [i,j] #row,column 
                                    
                                    
                    ## declaring regexes for searching and then function calls for the regexes
                    VTH_Pattern = re.compile('VTH', re.I)
                    VTH_ORIGIN = SearchXLWithRE(VTH_Pattern)
    
                    ChannelL_Pattern = re.compile('Channel Length', re.I)
                    ChannelL_ORIGIN = SearchXLWithRE(ChannelL_Pattern)
                    
                    Date_Pattern = re.compile('Date', re.I)
                    Date_ORIGIN = SearchXLWithRE(Date_Pattern)
                    
                    IONOFF_Pattern = re.compile('IONOFF', re.I)
                    IONOFF_ORIGIN = SearchXLWithRE(IONOFF_Pattern)
                    
                    VTO_Pattern = re.compile('VTO', re.I)
                    VTO_ORIGIN = SearchXLWithRE(VTO_Pattern)
            
                    # function defined for loop below to find all cell contents 
                    # based on an origin defined above
                    
                    def DataSeeker(ORIGIN,rowOffset,colOffset,X):
                        result_raw = ws.cell(row=ORIGIN[0]+rowOffset+r,
                                             column=ORIGIN[1]+colOffset).value
                        if not type(result_raw) == str:
                            result = X.format(result_raw)
                        else:
                            result = result_raw
                        return result

					def DataSeeker_noformat(ORIGIN,rowOffset,colOffset):
                        result = ws.cell(row=ORIGIN[0]+rowOffset+r,
                                             column=ORIGIN[1]+colOffset).value
                        return result
                    
                    
                    # for loop to return all data for the 4 subsites
                    # with offsets calculated from ORIGINS defined above
                    
                    for r in range(1,5):
                        test_date = ws.cell(row=Date_ORIGIN[0]+1,column=Date_ORIGIN[1]).value 
                        # no looping 1-4
                        ChannelL = DataSeeker(ChannelL_ORIGIN,0,0,"{:.2f}")
                        ChannelW = DataSeeker(ChannelL_ORIGIN,0,1,"{:.2f}")
                        MobilityMed = DataSeeker(ChannelL_ORIGIN,0,10,"{:.2f}")
                        MobSDev = DataSeeker(ChannelL_ORIGIN,0,4,"{:.2%}")
                        Ionoff = DataSeeker(IONOFF_ORIGIN,0,8,"{:.2e}")
                        VTO = DataSeeker_noformat(VTO_ORIGIN,0,0)    
                        VTH = DataSeeker(VTH_ORIGIN,0,0,"{:.2f}")
                        Yield = DataSeeker(ChannelL_ORIGIN,0,9,"{:.2%}")
                            
                        channel_data = (fname,test_date,ChannelL,ChannelW,MobilityMed,
                                        MobSDev,Ionoff,VTO,VTH,Yield)
                        
                        substrate_data.append(channel_data)
                        
                        writer.writerow(channel_data)
                       
                        print(fname,',',ChannelL,',',ChannelW,',',MobilityMed,',',MobSDev,
                              ',',Ionoff,',',VTO,',',VTH,',',Yield)
                        
                        continue

                    print('-$'*10,substrate)  ####### DON'T EDIT THIS VERSION !!!!!!!!!!
                    
        continue   

######## TIMING LOGIC #########

total_time = ((time.clock() - start_time))
print("--- %s seconds ---" % round(total_time,2))
print("%s seconds per file" % round(total_time/substrate))
###############################