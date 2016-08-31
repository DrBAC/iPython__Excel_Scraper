#######################################################################################################
#           this Python3 Program was written by Ben Coombs for use probing excel files                #
#              at CPI NetPark - Printable Electronics. It is not for general use                      #
#          and should be considered to be in beta-test mode. BAC Copyright March 2016.                #
#######################################################################################################
###########           ##############           #############         #################        #########

# new version featuring AS-YOU-GO WRITE OUT
# new version featuring LOCAL ORIGINS FOR BETTER SEARCHING
# searches for VTH, IONOFF and VTO locally now (directly indexed)
# searches now performed using REGEX
# formatting of output is now correct - incorporated in the function definition

import openpyxl
from openpyxl import load_workbook
from openpyxl.cell import get_column_letter, column_index_from_string
import time
import os
import pandas as pd
import csv
import re

#start clock to calculate timings
start_time = time.clock()
      
# initiate some variables/lists - substrate is index so can count number of files searched thru	  
substrate = 0                                               
outputfile = []
channel_data = []
substrate_data = []

print('Excel File Name',',','CL',',','CW',',','Mobility',',','SDev',',','On/Off',',','VTO',',',
      'VTH',',','Yield',',','Capacitance')

# file to write to is opened, named based on date
date = time.strftime('DIRECTORY TEST DATA OVERVIEW %Y-%m-%d.csv')#-%H%M%S)

with open(date, "w", newline='') as f:
    writer = csv.writer(f)
    writer.writerow(["File", "Date of test", "CL", "CW", "Mobility", "SDev", "On/Off", "VTO", "VTH", "Yield"])

    for root, dirs, files in os.walk(".", topdown=True):
        for fname in files:
            if fname.endswith('.xlsm') or fname.endswith('.xlsx'):
                wb = load_workbook(os.path.join(root, fname), data_only=True)
                names = wb.get_sheet_names()

                if 'validated Summary Data' in names:
                    ws = wb.get_sheet_by_name('validated Summary Data')

                    substrate = substrate + 1

                    Row_limit = ws.max_row
                    Column_limit = ws.max_column
                    Col_limit = get_column_letter(Column_limit)

                    def SearchXLWithRE(regex):
                        for i in range(1, Row_limit):
                            for j in range(1, Column_limit):
                                query = ws.cell(row = i, column = j).value
                                if type(query) == str:
                                    test = regex.search(query)
                                    if test: # checks if regex has found a match
                                        return [i,j] #row,column
                        return [100,100]
                                    
                    def DataSeeker(ORIGIN,rowOffset,colOffset,X):
                        result_raw = ws.cell(row=ORIGIN[0]+rowOffset+r,
                                                 column=ORIGIN[1]+colOffset).value
                        if not type(result_raw) == str:
                            result = X.format(result_raw)
                        else:
                            result = result_raw
                        return result
                    
                    def OriginFinder(Regex_Term):
                        Regex_Pattern = re.compile(Regex_Term, re.I)
                        Origin_Name = SearchXLWithRE(Regex_Pattern)
                        return Origin_Name

                    
                    # for loop to return all data for the 4 subsites
                    # with offsets calculated from ORIGINS defined during these calls (using Origin Finder func)
                    
                    for r in range(1,5):
                    # these first three don't actually loop as are single admin-y data bits
                        test_date = ws.cell(row=OriginFinder('Date')[0]+1,column=OriginFinder('Date')[1]).value
                        scientist = ws.cell(row=OriginFinder('Scientist')[0]+1,column=OriginFinder('Scientist')[1]).value
                       # Comments = ws.cell(row=OriginFinder('comment')[0],column=OriginFinder('comment')[1]+1).value
                        Capacitance = ws.cell(row=OriginFinder('Cap')[0]+1,column=OriginFinder('Cap')[1]).value
                        Temp1 = ws.cell(row=OriginFinder('Fisher')[0],column=OriginFinder('Fisher')[1]+1).value
                        Temp2 = ws.cell(row=OriginFinder('Calibrated')[0],column=OriginFinder('Calibrated')[1]+1).value
                        Hum1 = ws.cell(row=OriginFinder('Fisher')[0],column=OriginFinder('Fisher')[1]+2).value
                        Hum2 = ws.cell(row=OriginFinder('Calibrated')[0],column=OriginFinder('Calibrated')[1]+2).value
                        Ref = ws.cell(row=OriginFinder('Reference:')[0],column=OriginFinder('Reference:')[1]+1).value
                    # this lot that follow get looped
                        ChannelL = DataSeeker(OriginFinder('^Channel Length'),0,0,"{:.2f}")
                        ChannelW = DataSeeker(OriginFinder('^Channel Length'),0,1,"{:.2f}")
                        MobilityMed = DataSeeker(OriginFinder('^Channel Length'),0,10,"{:.2f}")
                        MobSDev = DataSeeker(OriginFinder('^Channel Length'),0,4,"{:.2%}")
                        Ionoff = DataSeeker(OriginFinder('^IONOFF'),0,8,"{:.2e}")
                        VTO = DataSeeker(OriginFinder('^VTO$'),0,0,"{:.2f}")    
                        VTH = DataSeeker(OriginFinder('^VTH'),0,0,"{:.2f}")
                        Yield = DataSeeker(OriginFinder('^Channel Length'),0,9,"{:.2%}")
                        channel_data = (fname,test_date,scientist,ChannelL,ChannelW,MobilityMed,
                                        MobSDev,Ionoff,VTO,VTH,Yield,Temp1,Temp2,Hum1,Hum2,Ref,Capacitance)
                        substrate_data.append(channel_data)
                        writer.writerow(channel_data)                       
                        print(fname,',',test_date,',',ChannelL,',',ChannelW,',',MobilityMed,','
                              ,MobSDev,',',Ionoff,',',VTO,',',VTH,',',Yield,',',Temp1,',',Temp2,',',Hum1,',',Hum2,',',Ref,',',Capacitance,',')                  
                        continue
                    print('-'*10,substrate)  ####### DON'T EDIT THIS VERSION !!!!!!!!!!
                    continue
        continue 

############## TIMING LOGIC ###############

total_time = ((time.clock() - start_time))
print("--- %s seconds ---" % round(total_time,2))
print("%s seconds per file" % round(total_time/substrate))

#######################################